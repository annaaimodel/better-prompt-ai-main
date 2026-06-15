#!/usr/bin/env python3
"""Single source of truth for the High-Ticket Sales / CSM directory.
Emits both an .xlsx spreadsheet and a self-contained searchable .html page."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# DATA  (edit here once -> rebuilds both outputs)
# Each section: (title, accent, [ (subsection_label, [ (name, desc, members, link) ... ]) ])
# ---------------------------------------------------------------------------
DATA = [
("Facebook", "#1877F2", [
 ("Job-specific - job boards / closer & setter opportunities", [
  ("High Ticket Closers and Setters \U0001F4A5\U0001F525 (JOB OPPORTUNITIES)", "Active job board for high-ticket closing & appointment-setting roles; posts 2-3x daily plus advice and events.", "~27,100", "https://www.facebook.com/groups/highticketsclosers/"),
  ("High-Ticket Closer Family™ (Members Only)", "Dan Lok's members-only community for graduates/closers - networking, opportunities and high-ticket closing support.", "Login required", "https://www.facebook.com/groups/highticketclosers/"),
  ("High Ticket Closer Opportunities", "Job-opportunity board connecting high-ticket closers with companies that are hiring.", "Login required", "https://www.facebook.com/groups/555294539537383/"),
  ("High Ticket Closers - NEW OPPORTUNITIES", "Posts new remote high-ticket closing roles and opportunities for setters/closers.", "Login required", "https://www.facebook.com/groups/highticketclosingopportunities/"),
  ("FREE JOB OPPORTUNITIES & HIGH TICKET SALES", "Free job-opportunity feed for high-ticket sales setters and closers.", "Login required", "https://www.facebook.com/groups/forhighticketclosers/"),
  ("High Ticket Inbound Closers Opportunity's", "Opportunity board focused on inbound high-ticket closing roles.", "Login required", "https://www.facebook.com/groups/393545198866970/"),
  ("High Ticket Sales & Closing | Closer Opportunities (Andrew Imbesi)", "Andrew Imbesi's group sharing high-ticket sales & closing job opportunities.", "Login required", "https://www.facebook.com/groups/andrewimbesihighticketsalesandclosingopportunities/"),
  ("High Ticket Sales - Closer Opportunities Job Board", "Dedicated job board listing high-ticket closer opportunities.", "Login required", "https://www.facebook.com/groups/highticketsalesca/"),
  ("Remote Sales Jobs | Appointment Setter and High Ticket (7-Figure Sales Pros)", "Remote sales job feed for appointment setters and high-ticket closers.", "Login required", "https://www.facebook.com/groups/7figuresalespros/"),
  ("High Ticket Closing: Sales Jobs For Setters & Closers", "Job board for high-ticket setter and closer roles.", "Login required", "https://www.facebook.com/groups/highticketsalesjobsforclosersandsetters/"),
  ("Matt's Remote Setters & Closers Network (High Ticket)", "Network posting remote high-ticket setter & closer roles plus a linked free Skool/cash competition.", "Login required", "https://www.facebook.com/groups/highticketcloserss/"),
  ("High Ticket Appointment Setters and Closers", "Opportunity group for high-ticket appointment setters and closers.", "Login required", "https://www.facebook.com/groups/267330199599984/"),
  ("High Ticket Appointment Setting | Closers & Sales Rep (Impact Closers)", "Job and opportunity board for high-ticket appointment setters, closers and sales reps.", "Login required", "https://www.facebook.com/groups/impactclosers/"),
  ("High Ticket Titans (Setters/Closers Opportunities)", "Opportunities and resources for high-ticket appointment setters and closers.", "Login required", "https://www.facebook.com/groups/1601202083962570/"),
  ("High Ticket SMMA / Appointment Setting & Closing \U0001F4B0", "Group for SMMA owners and high-ticket setters/closers - opportunities and tactics.", "Login required", "https://www.facebook.com/groups/702508090786804/"),
  ("CANADIAN JOBS - Appointment Setters & Sales Closers", "Canada-focused job board for appointment setters and sales closers.", "Login required", "https://www.facebook.com/groups/327324528755233/"),
  ("High Ticket Sales / Remote Closers - AUSTRALIAN", "Australia-focused group for high-ticket sales and remote closing roles.", "Login required", "https://www.facebook.com/groups/3470780863202640/"),
  ("Setter y Closer HIGH TICKET LATINOAMERICA", "Spanish-language Latin-America group for high-ticket setters and closers.", "Login required", "https://www.facebook.com/groups/settersyclosershighticket/"),
 ]),
 ("General - training, networking & community", [
  ("High-Ticket Sales - ELITE Community For Closers", "Community for closers focused on high-ticket sales skills, training and networking.", "Login required", "https://www.facebook.com/groups/eliteclosers/"),
  ("The Closers Dojo - High Ticket Sales Training & Job Opportunities", "Free remote high-ticket sales training plus job opportunities (Kris Cochrane).", "Login required", "https://www.facebook.com/groups/theclosersdojo/"),
  ("High-Ticket Closers | Remote Sales Opportunities & Training (The Advancing Mind)", "Blends remote sales opportunities with high-ticket closing training and mindset content.", "Login required", "https://www.facebook.com/groups/theadvancingmind/"),
  ("Sales People, High Ticket Closers & Setters Community", "General community for salespeople, high-ticket closers and setters to learn and connect.", "Login required", "https://www.facebook.com/groups/salespeopleandclosers/"),
  ("The Elite-Closers Community", "Community for elite closers sharing high-ticket sales strategy and support.", "Login required", "https://www.facebook.com/groups/elitesalesclosers/"),
  ("Club Closer Community | High Ticket Sales and Closers", "Community hub for high-ticket sales and closers.", "Login required", "https://www.facebook.com/groups/823218068832896/"),
  ("The Closers Community", "General community for closers around high-ticket sales.", "Login required", "https://www.facebook.com/groups/thecloserscom/"),
  ("Remote Closing Academy (Cole Gordon - private group)", "Private student community for Cole Gordon's Remote Closing Academy: training, livestreams and peer feedback.", "Login required", "https://www.facebook.com/groups/1754793088033322/"),
  ("Closer Babes - Remote High-Ticket Sales Training for Women", "Women-focused remote high-ticket sales training and community (group accessed via the program site).", "Via program site", "https://closerbabes.com/"),
 ]),
]),
("Skool", "#FFB800", [
 ("Job-specific - recruitment, placement & hiring marketplaces", [
  ("Remote Closing Community (RCC)", "Members-only marketplace connecting setters/closers with founders hiring; screened job posts every 24-48h. 7-day free trial then $27/mo.", "~1,000+", "https://www.skool.com/remote-closing-community-5761/about"),
  ("ORA - High Ticket Recruitment", "Free training + recruitment connecting 7-9 figure businesses with elite remote closers, setters and managers.", "Login required", "https://www.skool.com/free-high-ticket-sales-recruit-3019/about"),
  ("Sales Training & Placement (The Sales Agency)", "Robb Quinn's community: closer's playbook, weekly war-room calls and priority access to elite remote sales roles.", "Login required", "https://www.skool.com/thesalesagency/about"),
  ("Remote Sales Placement (The High Ticket Bridge)", "Trains and places high-ticket closers into remote sales roles.", "Login required", "https://www.skool.com/the-high-ticket-bridge/about"),
  ("High Ticket Sales | BOC™", "Free high-ticket sales network to hire and get hired into appointment-setting and closing roles.", "Login required", "https://www.skool.com/appointmentsetters/about"),
  ("High Ticket Setter™", "Positions itself as the #1 community training and creating top high-ticket appointment setters.", "Login required", "https://www.skool.com/high-ticket-setter-7793/about"),
  ("High Ticket Titans", "For appointment setters who want to perfect their skills and land jobs at top companies in their niche.", "Login required", "https://www.skool.com/high-ticket-titans-3792/about"),
  ("Superior Setters", "Community of remote-sales setters learning to become A-game appointment setters.", "Login required", "https://www.skool.com/superior-setters-1902/about"),
  ("Appointment Setter (Free)", "Coaching community (1,200+ setters coached) to book calls and land high-ticket opportunities.", "Login required", "https://www.skool.com/appointmentsetting/about"),
  ("Closers Inner Circle (Job Opportunities)", "Closer community posting remote high-ticket job opportunities.", "Login required", "https://www.skool.com/closers-circle/about"),
  ("Sales Community for Closers", "Community that regularly posts remote high-ticket closer roles with top commissions.", "Login required", "https://www.skool.com/sales-community-for-closers-1459"),
  ("Elite Closers", "Closer community that posts active high-ticket sales hiring ($5-8K/mo OTE roles).", "Login required", "https://www.skool.com/eliteclosers"),
  ("Staffing World", "Staffing/agency community recruiting remote high-ticket closers.", "Login required", "https://www.skool.com/agencyworld"),
  ("Elite Remote Closers Network", "Network for remote high-ticket closers with opportunities and frank industry guidance.", "Login required", "https://www.skool.com/elite-remote-closers-network-3228"),
  ("Remote Sales Job Accelerator", "Teaches how to land a remote sales role paying $4-10k/mo with no prior experience.", "Login required", "https://www.skool.com/remotesalesjobaccelerator/about"),
  ("Setter To Closer - Elite", "Community focused on progressing setters into closing roles.", "~78", "https://www.skool.com/setter-to-closer-elite-3793/about"),
 ]),
 ("General - training, mastermind & community", [
  ("High-Ticket Sales Society", "Free community (Chris Campbell) for mastering sales, marketing & mindset; learn to earn $10k-$25k+/mo as a remote closer.", "~2,200", "https://www.skool.com/high-ticket-sales/about"),
  ("High Ticket Force", "Community for closers, setters & online businesses focused on hiring, coaching, leadership and scaling revenue.", "Login required", "https://www.skool.com/high-ticket-force-3713/about"),
  ("Skool for High Ticket Sales (High Ticket Millions)", "For experts who've sold $10k+ of their own offers; win higher-paying clients and recurring revenue via Skool.", "Login required", "https://www.skool.com/high-ticket-millions-6222/about"),
  ("High Ticket Sales Blueprint", "Mission-driven community empowering high-ticket closers to master sales, accountability and financial freedom.", "Login required", "https://www.skool.com/traveling-sales-team-9554"),
  ("High Ticket Sales Training", "Free group for high-ticket setters and closers looking to reach their A-player potential.", "Login required", "https://www.skool.com/high-ticket-sales-training/about"),
  ("High Ticket Sales Battle", "Free group for remote salespeople wanting to fast-track into remote high-ticket sales from home.", "Login required", "https://www.skool.com/high-ticket-sales-battle"),
  ("The High Ticket Sales Academy (Psychological Sales)", "Sales-skills community focused on the psychology of closing high-ticket offers.", "Login required", "https://www.skool.com/psychological-sales"),
  ("The Serial Sales Community", "General sales community covering high-ticket sales skills and growth.", "Login required", "https://www.skool.com/serialsales/about"),
  ("Remote Closing Bootcamp", "Free community to break into remote high-ticket sales - where to find offers and the fundamentals of closing.", "Login required", "https://www.skool.com/closing-authority-saless/about"),
  ("XCloser.com (Remote Closing)", "Free remote-sales community with 101 coaching calls on your setter/closer journey.", "Login required", "https://www.skool.com/xclosercom-free-training-3117"),
  ("The Closers Academy", "Mastermind for ambitious remote high-ticket closers, coaches and experts (CTA system).", "Login required", "https://www.skool.com/theclosersacademy/about"),
  ("High Ticket Sales University (Remote Closer Mastery)", "Training community for remote high-ticket closing skills.", "Login required", "https://www.skool.com/remote-closer-mastery-3732/about"),
  ("The Remote Closing Lab", "Serves remote high-ticket closers, SDRs and setters with coaching, recruitment and accountability.", "Login required", "https://www.skool.com/remoteclosinglab/about"),
  ("Quantum Sales Academy™", "Sales academy community training remote high-ticket closers (also posts roles).", "Login required", "https://www.skool.com/qsa"),
  ("Sales Genius University", "Training community running webinars on breaking into remote high-ticket sales.", "Login required", "https://www.skool.com/salesgeniuses"),
  ("Remote Closing Club", "Community for remote closers focused on high-ticket sales skills.", "Login required", "https://www.skool.com/remote-closing-club-4755/about"),
  ("High Ticket Closers", "General high-ticket closer community on Skool.", "Login required", "https://www.skool.com/high-ticket-closers-4841/about"),
 ]),
]),
("Discord", "#5865F2", [
 ("Job-specific - offer placement / job & lead boards", [
  ("High Flow Closers / FlowClose", "Bills itself as the #1 sales & offer-placement community - free offer placements, scripts, networking and weekly giveaways for closers/setters.", "Join to view", "https://top.gg/discord/servers/766313238700785664"),
  ("High Ticket Sales Closers", "Offers ~20,000 free email/phone leads per month, free high-ticket courses, networking and webinars.", "Join to view", "https://discord.me/highticketsalesclosers"),
  ("Appointment Setter Life", "Setter-focused server with free leads and high-ticket content for setters/closers.", "Join to view", "https://discord.me/appsetterlife"),
  ("Appointment Setter Jobs", "DISBOARD-listed server functioning as an appointment-setter job board.", "Join to view", "https://disboard.org/server/1332654195958087681"),
 ]),
 ("General - training, networking & community", [
  ("Prosper", "Results-driven community for mastering high-ticket sales and scaling income - frameworks, closing, client acquisition, personal branding.", "Join to view", "https://disboard.org/servers/tag/high-ticket-sales"),
  ("TheCloserCommunity", "100% free community where high-ticket closers motivate each other to close more deals; shares leads and courses.", "Join to view", "https://discord.me/8qxyr"),
  ("High Ticket Sales Rep", "Server for high-ticket sales reps - free leads, courses and networking.", "Join to view", "https://discord.me/highticketsalesrep"),
  ("Sales Pros", "General sales community covering high-ticket closing skills and networking.", "Join to view", "https://discord.me/salespros"),
  ("DISBOARD - 'high-ticket-sales' tag (directory)", "Browse/discovery page listing many high-ticket-sales Discord servers, sorted by activity.", "Directory", "https://disboard.org/servers/tag/high-ticket-sales"),
 ]),
]),
("LinkedIn", "#0A66C2", [
 ("Job-specific - recruiters / pages posting high-ticket roles", [
  ("Elite Closers (company page)", "High-ticket sales company connecting the top 1% of business owners with top 1% sales professionals; posts roles.", "Followers vary", "https://www.linkedin.com/company/eliteclosers"),
  ("High Ticket Closer Club", "Invitation-only sales-training/opportunity built around Grant Cardone content for closers raising their income.", "Followers vary", "https://www.linkedin.com/company/high-ticket-closer-club"),
  ("LinkedIn Jobs - 'high ticket closer / setter'", "LinkedIn job-search feed where agencies (WFS Group, 1 Call Closers, Grant Cardone Enterprises, etc.) post high-ticket roles.", "N/A (job board)", "https://www.linkedin.com/jobs/search/?keywords=high%20ticket%20closer"),
 ]),
 ("General - training pages & professional community", [
  ("High Ticket Sales Academy", "Company page sharing high-ticket sales professional-development content and community.", "Followers vary", "https://www.linkedin.com/company/high-ticket-sales-academy"),
  ("Note on LinkedIn Groups", "LinkedIn has largely sunset native 'Groups' for this niche - presence now lives in company pages, hashtags (#highticketsales, #remoteclosing) and the jobs feed.", "-", ""),
 ]),
]),
("Customer Success / Client Success Manager (CSM) - all platforms", "#34A853", [
 ("Job-specific - high-ticket / remote CSM training & placement", [
  ("Client Success Club (CSC) - Robert Lyon  [Skool]", "Bills itself as the #1 community for landing high-paying remote CSM roles ($10k-25k/mo) for high-ticket businesses; training, job board, weekly coaching.", "Login required", "https://www.skool.com/client-success/about"),
  ("Client Success Club (alt)  [Skool]", "Client-success-manager training & community focused on landing remote high-ticket CSM roles.", "Login required", "https://www.skool.com/client-success-club/about"),
  ("Remote Sales / CSM Jobs (Sales Pipeline Pros - Group #3)  [Facebook]", "Facebook job board posting remote sales AND customer/client success manager roles.", "Login required", "https://www.facebook.com/groups/648104543815634/"),
  ("Virtual Assistant - Job Offers board  [Skool]", "Active job board that regularly posts remote Client/Customer Success Manager openings alongside VA roles.", "Login required", "https://www.skool.com/virtual-assistant-8842"),
 ]),
 ("General - customer-success professional communities", [
  ("Customer Success Collective  [LinkedIn + Slack]", "Global CS community (junior CSM to CCO) with a 10,000+ member Slack, templates, articles, events and job board.", "10,000+ (Slack)", "https://www.customersuccesscollective.com/customer-success-community/"),
  ("Customer Success Network (CSN)  [LinkedIn + Workplace]", "First global peer-learning CS community; free access to an online community to ask questions and network with CSMs/CS leaders.", "Free to join", "https://customersuccess.network/"),
  ("The Customer Success Forum (CS Association)  [Web/LinkedIn]", "Long-running professional discussion forum spanning CEOs to practising CSMs across all company sizes worldwide.", "~65,000+", "https://www.customersuccessassociation.com/forum/"),
  ("Customer Success Leadership Network  [LinkedIn]", "Content + connection network focused on sustaining the CS function and CS leadership.", "~14,300 followers", "https://www.linkedin.com/company/cs-leadership-network"),
  ("Gain Grow Retain  [Slack/Web]", "Peer-to-peer community for B2B SaaS customer-success leaders - team structure, retention, cross-collaboration.", "Free to join", "https://gaingrowretain.com/"),
  ("CS Cafe  [Slack]", "Slack network for customer-success professionals to connect, share and find roles.", "Free to join", "https://www.thecscafe.com/p/community"),
  ("Customer Success Network (CSM Practice)  [Facebook]", "Facebook presence of the CS community - templates, videos, reviews and event announcements.", "Login required", "https://www.facebook.com/customersuccessnetwork"),
 ]),
]),
]

NOTE = ("Member counts shown are those publicly visible at time of research (12 Jun 2026). "
        "'Login required' / 'Join to view' means the platform hides the count behind a login. "
        "Discord servers are reached via invite/listing pages and counts fluctuate. "
        "LinkedIn has largely retired native Groups for this niche, so its presence is company pages + the jobs feed. "
        "CSM rows are tagged with their platform in [brackets]. Many of these are marketing/lead-gen communities for paid programs - verify before outreach.")

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def build_xlsx(path):
    wb = Workbook(); ws = wb.active; ws.title = "High Ticket Groups"
    PFONT = Font(bold=True, color="FFFFFF", size=14)
    SUB_FILL = PatternFill("solid", fgColor="2E75B6"); SUB_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEAD_FILL = PatternFill("solid", fgColor="D9E1F2"); HEAD_FONT = Font(bold=True, color="1F3864")
    WRAP = Alignment(wrap_text=True, vertical="top"); TOPA = Alignment(vertical="top")
    thin = Side(style="thin", color="BFBFBF"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    r = 1
    def hexc(c): return c.replace("#", "").upper()
    for title, accent, subs in DATA:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title.upper())
        c.fill = PatternFill("solid", fgColor=hexc(accent)); c.font = PFONT
        c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[r].height = 26; r += 1
        for label, rows in subs:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=label); c.fill = SUB_FILL; c.font = SUB_FONT
            c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[r].height = 20; r += 1
            for i, h in enumerate(["Group / Community", "What it's advertised as", "Members", "Link"], 1):
                cc = ws.cell(row=r, column=i, value=h); cc.fill = HEAD_FILL; cc.font = HEAD_FONT; cc.border = BORDER
            r += 1
            for name, desc, members, link in rows:
                for i, v in enumerate([name, desc, members, link], 1):
                    cc = ws.cell(row=r, column=i, value=v); cc.border = BORDER
                    cc.alignment = WRAP if i in (1, 2) else TOPA
                lc = ws.cell(row=r, column=4)
                if link.startswith("http"):
                    lc.hyperlink = link; lc.font = Font(color="0563C1", underline="single")
                r += 1
            r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    n = ws.cell(row=r, column=1, value="Notes: " + NOTE)
    n.font = Font(italic=True, size=9, color="595959"); n.alignment = WRAP; ws.row_dimensions[r].height = 70
    for col, w in {"A": 50, "B": 66, "C": 17, "D": 60}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"; wb.save(path); print("Saved", path)

# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------
def build_html(path):
    total = sum(len(rows) for _, _, subs in DATA for _, rows in subs)
    sections_json = []
    for title, accent, subs in DATA:
        sj = {"title": title, "accent": accent, "subs": []}
        for label, rows in subs:
            kind = "job" if label.lower().startswith("job") else "general"
            sj["subs"].append({"label": label, "kind": kind,
                "rows": [{"name": n, "desc": d, "members": m, "link": l} for n, d, m, l in rows]})
        sections_json.append(sj)
    data_js = json.dumps(sections_json, ensure_ascii=False)
    note_js = json.dumps(NOTE, ensure_ascii=False)
    doc = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>High-Ticket Sales & CSM Communities Directory</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@500;600;700;800&display=swap');
:root{--bg:#0a0a0b;--card:#141312;--card2:#1a1714;--line:rgba(201,156,56,.14);--line2:#262320;--txt:#f3efe6;--mut:#9a948a;--accent:#c99c38;--gold:#c99c38;--gold2:#f2dd88;}
*{box-sizing:border-box}
body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;background:var(--bg);color:var(--txt);line-height:1.45}
header{padding:32px 20px 18px;text-align:center;background:radial-gradient(900px 300px at 50% -40%,#22304d,transparent)}
h1{margin:0 0 6px;font-size:26px;letter-spacing:.2px}
.sub{color:var(--mut);font-size:14px;margin:0}
.nav{text-align:center;margin:12px 0 0;display:flex;gap:8px;justify-content:center;flex-wrap:wrap}
.nav a{color:var(--mut);font-size:13px;text-decoration:none;border:1px solid var(--line);padding:5px 11px;border-radius:8px}
.nav a:hover{border-color:var(--accent);color:var(--txt)}.nav a.on{color:#0a0a0b;background:linear-gradient(175deg,#f6e29a 0%,#d9b24c 45%,#b8862f 100%);border-color:var(--accent)}
.wrap{max-width:1100px;margin:0 auto;padding:0 16px 60px}
.controls{position:sticky;top:0;z-index:5;background:rgba(15,20,32,.92);backdrop-filter:blur(8px);padding:14px 0;margin-bottom:6px;border-bottom:1px solid var(--line)}
.row{display:flex;gap:10px;flex-wrap:wrap;align-items:center;justify-content:center}
input[type=search]{flex:1;min-width:240px;max-width:460px;padding:11px 14px;border-radius:10px;border:1px solid var(--line);background:var(--card);color:var(--txt);font-size:15px;outline:none}
input[type=search]:focus{border-color:var(--accent)}
.chip{padding:8px 13px;border-radius:999px;border:1px solid var(--line);background:var(--card);color:var(--mut);cursor:pointer;font-size:13px;font-weight:600;transition:.15s}
.chip:hover{color:var(--txt)}
.chip.on{background:var(--accent);border-color:var(--accent);color:#fff}
.count{color:var(--mut);font-size:13px;text-align:center;margin:8px 0 2px}
section.platform{margin-top:26px}
.phead{display:flex;align-items:center;gap:10px;font-size:19px;font-weight:700;padding:9px 14px;border-radius:10px;color:#fff}
.subhead{margin:16px 0 8px;font-size:13px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;color:var(--mut);display:flex;align-items:center;gap:8px}
.tag{font-size:10px;padding:2px 8px;border-radius:999px;font-weight:700;letter-spacing:.3px}
.tag.job{background:#3a2a14;color:#ffce85}
.tag.general{background:#15303a;color:#7fd7ff}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:12px}
.card{background:var(--card);border:1px solid var(--line);border-radius:13px;padding:14px 15px;display:flex;flex-direction:column;gap:7px;transition:.15s}
.card:hover{border-color:var(--accent);transform:translateY(-2px)}
.cname{font-weight:700;font-size:15px;line-height:1.3}
.cdesc{color:var(--mut);font-size:13px;flex:1}
.cfoot{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-top:2px}
.mem{font-size:12px;color:#cbd6ea;background:var(--card2);padding:3px 9px;border-radius:999px;white-space:nowrap}
a.visit{font-size:13px;font-weight:700;color:#0a0a0b;background:linear-gradient(175deg,#f6e29a 0%,#d9b24c 45%,#b8862f 100%);padding:7px 12px;border-radius:8px;text-decoration:none;white-space:nowrap}
a.visit:hover{filter:brightness(1.12)}
.nolink{font-size:12px;color:var(--mut)}
.note{margin-top:30px;font-size:12px;color:var(--mut);background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px 16px}
.empty{text-align:center;color:var(--mut);padding:50px 0;display:none}
mark{background:#5b8cff44;color:inherit;border-radius:3px}
footer{text-align:center;color:var(--mut);font-size:12px;padding:24px}
</style></head>
<body>
<header>
<h1>High-Ticket Sales &amp; CSM Communities</h1>
<p class="sub">A directory across Facebook, Skool, Discord &amp; LinkedIn &middot; __TOTAL__ communities &middot; researched 12 Jun 2026</p>
<div class="nav"><a href="/jobs">Fresh today</a><a href="/jobs/all">The Vault</a><a href="/jobs/boards">&starf; Boards</a><a href="/cv">CV match</a><a class="on" href="/directory">Communities</a></div>
</header>
<div class="wrap">
<div class="controls">
<div class="row">
<input type="search" id="q" placeholder="Search name or description (e.g. setter, free, women, CSM)...">
</div>
<div class="row" id="platChips" style="margin-top:10px"></div>
<div class="row" id="kindChips" style="margin-top:8px">
<span class="chip kind on" data-kind="all">All types</span>
<span class="chip kind" data-kind="job">Job-specific</span>
<span class="chip kind" data-kind="general">General</span>
</div>
<div class="count" id="count"></div>
</div>
<div id="results"></div>
<div class="empty" id="empty">No communities match your search.</div>
<div class="note" id="note"></div>
</div>
<footer>Built as a companion to the spreadsheet &middot; links &amp; member counts change frequently &mdash; verify before outreach.</footer>
<script>
const DATA = __DATA__;
const NOTE = __NOTE__;
document.getElementById('note').textContent = "Notes: " + NOTE;
let state = {q:"", plat:"all", kind:"all"};

const platChips = document.getElementById('platChips');
platChips.innerHTML = '<span class="chip plat on" data-plat="all">All platforms</span>' +
  DATA.map(s=>`<span class="chip plat" data-plat="${s.title}" style="--a:${s.accent}">${s.title.split(' - ')[0].split(' / ')[0].split(' (')[0]}</span>`).join('');

function esc(t){return t.replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}
function hi(t){if(!state.q)return esc(t);const i=t.toLowerCase().indexOf(state.q.toLowerCase());if(i<0)return esc(t);return esc(t.slice(0,i))+'<mark>'+esc(t.slice(i,i+state.q.length))+'</mark>'+esc(t.slice(i+state.q.length));}

function render(){
  const res=document.getElementById('results');res.innerHTML='';let shown=0;
  for(const s of DATA){
    if(state.plat!=='all' && s.title!==state.plat) continue;
    let secHtml='';
    for(const sub of s.subs){
      if(state.kind!=='all' && sub.kind!==state.kind) continue;
      const rows=sub.rows.filter(r=>{
        if(!state.q) return true;
        const q=state.q.toLowerCase();
        return r.name.toLowerCase().includes(q)||r.desc.toLowerCase().includes(q);
      });
      if(!rows.length) continue;
      secHtml+=`<div class="subhead"><span class="tag ${sub.kind}">${sub.kind==='job'?'JOB-SPECIFIC':'GENERAL'}</span>${esc(sub.label)}</div><div class="grid">`;
      for(const r of rows){
        shown++;
        const link=r.link? `<a class="visit" href="${r.link}" target="_blank" rel="noopener">Visit &rarr;</a>` : `<span class="nolink">no direct link</span>`;
        secHtml+=`<div class="card"><div class="cname">${hi(r.name)}</div><div class="cdesc">${hi(r.desc)}</div><div class="cfoot"><span class="mem">${esc(r.members)}</span>${link}</div></div>`;
      }
      secHtml+='</div>';
    }
    if(secHtml){
      const sec=document.createElement('section');sec.className='platform';
      sec.innerHTML=`<div class="phead" style="background:${s.accent}">${esc(s.title)}</div>`+secHtml;
      res.appendChild(sec);
    }
  }
  document.getElementById('empty').style.display=shown?'none':'block';
  document.getElementById('count').textContent=shown+' communit'+(shown===1?'y':'ies')+' shown';
}
document.getElementById('q').addEventListener('input',e=>{state.q=e.target.value.trim();render();});
platChips.addEventListener('click',e=>{if(!e.target.classList.contains('plat'))return;document.querySelectorAll('.plat').forEach(c=>c.classList.remove('on'));e.target.classList.add('on');state.plat=e.target.dataset.plat;render();});
document.getElementById('kindChips').addEventListener('click',e=>{if(!e.target.classList.contains('kind'))return;document.querySelectorAll('.kind').forEach(c=>c.classList.remove('on'));e.target.classList.add('on');state.kind=e.target.dataset.kind;render();});
render();
</script>
</body></html>"""
    doc = doc.replace("__DATA__", data_js).replace("__NOTE__", note_js).replace("__TOTAL__", str(total))
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)
    print("Saved", path, "(", total, "communities )")

if __name__ == "__main__":
    build_xlsx("/home/user/better-prompt-ai-main/High_Ticket_Sales_Groups.xlsx")
    build_html("/home/user/better-prompt-ai-main/high_ticket_directory.html")
